from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import TemplateView, ListView, DetailView
from django.views import View
from django.db.models import Count, Avg, Max, Q
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.contrib import messages
from .decorators import admin_required_method, login_required_method
from .models import Equipo, HistorialDisponibilidad, ConfiguracionGlobal, Marca, TipoEquipo, UserProfile
from .forms import EquipoImportForm, UserProfileForm, ConfiguracionGlobalForm, MarcaForm, TipoEquipoForm, EquipoForm, PasswordChangeForm
import csv
from django.http import HttpResponse

@login_required_method
class DashboardView(TemplateView):
    template_name = 'monitor/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.db.models import Count, Q, Avg
        from django.utils import timezone
        import datetime
        import json
        from django.core.serializers.json import DjangoJSONEncoder
        
        # 1. Top 4 Brands Stats
        # We need: Brand Name, Total Count, Down Count, Status Color (implicit)
        top_brands = Equipo.objects.values('marca__nombre', 'marca__color').annotate(
            total=Count('id'),
            down=Count('id', filter=Q(is_online=False))
        ).order_by('-total')[:4]
        
        context['brand_stats'] = top_brands

        # 2. Network Latency Chart (Last 24h Average)
        now = timezone.now()
        start_24h = now - datetime.timedelta(hours=24)
        
        # Aggregate latency by hour or simpler chunks. 
        # For simplicity/speed in this prototype, we'll fetch all history in 24h and aggregate in python 
        # (Note: In production with millions of rows, use database truncation/group by)
        
        from django.db.models.functions import TruncHour
        from django.db.models import DateTimeField

        # 3. Network Latency Chart (Last 24h Average - Hourly)
        current_tz = timezone.get_current_timezone()
        now = timezone.now()
        start_24h = now - datetime.timedelta(hours=24)
        
        latency_qs = HistorialDisponibilidad.objects.filter(
            timestamp__gte=start_24h,
            estado='ONLINE'
        ).annotate(
            hour=TruncHour('timestamp', output_field=DateTimeField(), tzinfo=current_tz)
        ).values('hour').annotate(
            avg_latency=Avg('latencia_ms')
        ).order_by('hour')
        
        latency_labels = []
        latency_values = []
        for item in latency_qs:
            # Force server-side formatted string labels
            label = item['hour'].strftime('%H:%M')
            val = round(item['avg_latency'], 1)
            latency_labels.append(label)
            latency_values.append(val)
            
        context['latency_labels'] = json.dumps(latency_labels, cls=DjangoJSONEncoder)
        context['latency_data'] = json.dumps(latency_values, cls=DjangoJSONEncoder)
        
        # 3. Packet Loss
        avg_packet_loss = HistorialDisponibilidad.objects.filter(
            timestamp__gte=start_24h
        ).aggregate(avg=Avg('packet_loss'))['avg'] or 0
        context['packet_loss'] = round(avg_packet_loss, 2)

        # 4. Offline Devices List (Enriched)
        # We need downtime duration. 
        offline_devices = []
        raw_offline = Equipo.objects.filter(is_online=False, estado='ACTIVO').select_related('marca')
        for dev in raw_offline:
            downtime = "N/A"
            if dev.last_seen:
                diff = now - dev.last_seen
                # Format: HH:MM:SS
                total_seconds = int(diff.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                downtime = f"{hours:02}:{minutes:02}:{seconds:02}"
            
            offline_devices.append({
                'id_equipo': dev.id_equipo,
                'ip': dev.ip,
                'marca': dev.marca.nombre if dev.marca else 'Desconocido',
                'downtime': downtime,
                'id': dev.id # for url
            })
            
        context['offline_list'] = offline_devices[:5] # Top 5
        context['total_monitored'] = Equipo.objects.count()
        context['total_down'] = len(raw_offline)
        
        # Billing events for today and tomorrow
        from datetime import date, timedelta
        from .models import EventoFacturacion
        
        today = date.today()
        tomorrow = today + timedelta(days=1)
        
        # Get FACTURACION events for today
        eventos_hoy = EventoFacturacion.objects.filter(
            fecha=today,
            tipo_evento='FACTURACION'
        ).select_related('porcion').order_by('porcion__nombre')
        
        # Get FACTURACION events for tomorrow
        eventos_manana = EventoFacturacion.objects.filter(
            fecha=tomorrow,
            tipo_evento='FACTURACION'
        ).select_related('porcion').order_by('porcion__nombre')
        
        context['eventos_hoy'] = list(eventos_hoy)
        context['eventos_manana'] = list(eventos_manana)

        return context

@login_required_method
class EquipoListView(ListView):
    model = Equipo
    template_name = 'monitor/equipo_list.html'
    context_object_name = 'equipos'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['marcas'] = Marca.objects.all()
        context['tipos'] = TipoEquipo.objects.all() # For filter
        context['total_active'] = Equipo.objects.filter(estado='ACTIVO').count()
        return context

    def get_queryset(self):
        from django.db.models import Subquery, OuterRef, IntegerField
        from django.db.models.functions import Cast
        
        # Subquery for latest latency
        latest_latency = HistorialDisponibilidad.objects.filter(
            equipo=OuterRef('pk')
        ).order_by('-timestamp').values('latencia_ms')[:1]

        qs = super().get_queryset().annotate(
            last_latency=Cast(Subquery(latest_latency), output_field=IntegerField())
        ).select_related('marca', 'tipo').order_by('id_equipo')
        
        # Filtering
        query = self.request.GET.get('q')
        if query:
            qs = qs.filter(ip__icontains=query) | qs.filter(id_equipo__icontains=query)
            
        estado = self.request.GET.get('estado')
        if estado:
            qs = qs.filter(estado=estado)
            
        marca_id = self.request.GET.get('marca')
        if marca_id:
            qs = qs.filter(marca_id=marca_id)

        tipo_id = self.request.GET.get('tipo')
        if tipo_id:
            qs = qs.filter(tipo_id=tipo_id)

        medio = self.request.GET.get('medio')
        if medio:
            qs = qs.filter(medio_comunicacion=medio)

        comunicacion = self.request.GET.get('comunicacion')
        if comunicacion:
            if comunicacion == 'ONLINE':
                qs = qs.filter(is_online=True)
            elif comunicacion == 'OFFLINE':
                qs = qs.filter(is_online=False)
            
        return qs
        
    def get_template_names(self):
        if self.request.htmx:
            return ['monitor/partials/equipo_list_rows.html']
        return ['monitor/equipo_list.html']

class GlobalSearchView(ListView):
    model = Equipo
    template_name = 'monitor/partials/search_results.html'
    context_object_name = 'results'
    paginate_by = 5

    def get_queryset(self):
        query = self.request.GET.get('q', '')
        if len(query) < 2:
            return Equipo.objects.none()
        
        return Equipo.objects.filter(
            Q(id_equipo__icontains=query) | 
            Q(ip__icontains=query) |
            Q(marca__nombre__icontains=query)
        ).select_related('marca')[:5]


@login_required_method
class EquipoDetailView(DetailView):
    model = Equipo
    template_name = 'monitor/equipo_detail.html'
    context_object_name = 'equipo'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        import json
        from django.core.serializers.json import DjangoJSONEncoder
        from django.utils import timezone
        import datetime
        
        now = timezone.now()
        start_time = now - datetime.timedelta(hours=24)
        
        history = HistorialDisponibilidad.objects.filter(
            equipo=self.object,
            timestamp__gte=start_time
        ).values('timestamp', 'latencia_ms', 'estado').order_by('timestamp')
        
        # Prepare data for ApexCharts: Split into labels (server-time) and values
        # Value: 1 for ONLINE, 0 for OFFLINE/TIMEOUT
        chart_labels = []
        chart_values = []
        
        current_tz = timezone.get_current_timezone()
        
        for h in history:
            local_dt = h['timestamp'].astimezone(current_tz)
            chart_labels.append(local_dt.strftime('%d/%m %H:%M'))
            chart_values.append(1 if h['estado'] == 'ONLINE' else 0)
        
        context['chart_labels'] = chart_labels
        context['chart_values'] = chart_values
        
        # Calculate availability for the period (48h)
        total_checks = len(history)
        online_checks = sum(1 for h in history if h['estado'] == 'ONLINE')
        availability = round((online_checks / total_checks * 100), 2) if total_checks > 0 else 0
        context['availability'] = availability
        
        return context

from django.db.models import Count, Q, F, ExpressionWrapper, FloatField
from django.utils import timezone
import datetime

class ReporteView(ListView):
    model = Equipo
    template_name = 'monitor/reportes.html'
    context_object_name = 'equipos'
    paginate_by = 10

    def get_queryset(self):
        now = timezone.now()
        
        # Date Range Filtering
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        
        if start_date_str and end_date_str:
            try:
                # Use current active timezone (America/Lima) to parse expected range
                current_tz = timezone.get_current_timezone()
                
                # Start of day
                naive_start = datetime.datetime.strptime(start_date_str, '%Y-%m-%d')
                start_date = timezone.make_aware(naive_start, current_tz)
                
                # End of day
                naive_end = datetime.datetime.strptime(end_date_str, '%Y-%m-%d') + datetime.timedelta(days=1) - datetime.timedelta(microseconds=1)
                end_date = timezone.make_aware(naive_end, current_tz)
            except ValueError:
                start_date = now - datetime.timedelta(days=30)
                end_date = now
        else:
            start_date = now - datetime.timedelta(days=30)
            end_date = now

        # Calculate availability based on Ping History in Range
        qs = Equipo.objects.filter(estado='ACTIVO').annotate(
            total_checks=Count('historial', filter=Q(historial__timestamp__gte=start_date, historial__timestamp__lte=end_date)),
            online_checks=Count('historial', filter=Q(historial__timestamp__gte=start_date, historial__timestamp__lte=end_date, historial__estado='ONLINE'))
        )
        
        # Filtering
        query = self.request.GET.get('q')
        if query:
            qs = qs.filter(Q(ip__icontains=query) | Q(id_equipo__icontains=query))
            
        marca_id = self.request.GET.get('marca')
        if marca_id:
            qs = qs.filter(marca_id=marca_id)

        estado = self.request.GET.get('estado')
        if estado:
            qs = qs.filter(estado=estado)
            
        qs = qs.order_by('id_equipo')
        return qs

        return context

    def get_context_data(self, **kwargs):
        # This wrapper calls the merged logic if we were keeping it, 
        # but since we are REPLACING the file content in chunks, 
        # I need to be careful. The user tool `replace_file_content` replaces a contiguous block.
        # The two methods are adjacent? No, one ends at 289, the other starts at 291.
        # So I can replace the whole block from 245 to the end of the second method.
        pass

# ... Wait, I should write the actual python code for the replacement content.

# Merged get_context_data for ReporteView
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Current filter values
        try:
            current_marca_id = int(self.request.GET.get('marca', 0))
        except (ValueError, TypeError):
            current_marca_id = 0
            
        current_estado = self.request.GET.get('estado', '')

        # Prepare Marcas with selected flag
        marcas_list = []
        all_marcas = Marca.objects.all()
        for m in all_marcas:
            marcas_list.append({
                'id': m.id,
                'nombre': m.nombre,
                'selected': m.id == current_marca_id
            })
        context['marcas_list'] = marcas_list

        # Prepare Estados with selected flag
        estados_list = [
            {'value': '', 'label': 'Todos', 'selected': current_estado == ''},
            {'value': 'ACTIVO', 'label': 'Activo', 'selected': current_estado == 'ACTIVO'},
            {'value': 'INACTIVO', 'label': 'Inactivo', 'selected': current_estado == 'INACTIVO'},
        ]
        context['estados_list'] = estados_list
        
        # Pass back dates for form
        start_date_str = self.request.GET.get('start_date', '')
        end_date_str = self.request.GET.get('end_date', '')
        context['start_date'] = start_date_str
        context['end_date'] = end_date_str

        # Pre-calculate percentage for template
        for equipo in context['equipos']:
            if equipo.total_checks > 0:
                equipo.availability = round((equipo.online_checks / equipo.total_checks) * 100, 1)
                equipo.downtime_count = equipo.total_checks - equipo.online_checks
            else:
                equipo.availability = 0
                equipo.downtime_count = 0

        # --- MERGED DASHBOARD STATS (from duplicate method) ---
        from django.utils import timezone
        import datetime
        from django.db.models import Count, Avg
        
        # Recalculate range for global stats if needed, or use default last 24h
        now = timezone.now()
        last_24h = now - datetime.timedelta(hours=24)
        last_7d = now - datetime.timedelta(days=7)

        # 1. Total Outages (Last 24h)
        total_outages_24h = HistorialDisponibilidad.objects.filter(
            timestamp__gte=last_24h,
            estado='OFFLINE'
        ).count()
        context['total_outages_24h'] = total_outages_24h

        # 2. Global Average Latency (Last 24h)
        avg_latency_24h = HistorialDisponibilidad.objects.filter(
            timestamp__gte=last_24h,
            latencia_ms__isnull=False
        ).aggregate(Avg('latencia_ms'))['latencia_ms__avg']
        context['avg_latency_24h'] = round(avg_latency_24h or 0, 1)

        # 3. Worst Performing Devices (Most Offline events in last 7d)
        worst_devices = Equipo.objects.filter(
            historial__timestamp__gte=last_7d,
            historial__estado='OFFLINE'
        ).annotate(
            offline_count=Count('historial')
        ).order_by('-offline_count')[:5]
        context['worst_devices'] = worst_devices

        return context

class ReporteIndividualView(TemplateView):
    template_name = 'monitor/reporte_individual.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.utils import timezone
        import datetime
        from django.db.models import Count, Q
        
        # 1. Handle Device Selection by String ID (id_equipo)
        equipo_code = self.request.GET.get('equipo_code')
        
        # List for search dropdown (include brand)
        context['equipos_search'] = Equipo.objects.filter(estado='ACTIVO').values('id', 'id_equipo', 'ip', 'marca__nombre').order_by('id_equipo')
        
        if equipo_code:
            # Try to find by id_equipo, or fall back to PK if it's numeric (legacy support)
            equipo = None
            try:
                equipo = Equipo.objects.get(id_equipo=equipo_code)
            except Equipo.DoesNotExist:
                # Fallback: maybe passed PK?
                if str(equipo_code).isdigit():
                    try:
                        equipo = Equipo.objects.get(pk=equipo_code)
                    except Equipo.DoesNotExist:
                        pass
            
            
            if equipo:
                context['selected_equipo'] = equipo
            
            # 2. Parse Dates (Reuse logic or keep simple default 30d)
            start_date_str = self.request.GET.get('start_date')
            end_date_str = self.request.GET.get('end_date')
            now = timezone.now()
            
            if start_date_str and end_date_str:
                try:
                    current_tz = timezone.get_current_timezone()
                    naive_start = datetime.datetime.strptime(start_date_str, '%Y-%m-%d')
                    start_date = timezone.make_aware(naive_start, current_tz)
                    naive_end = datetime.datetime.strptime(end_date_str, '%Y-%m-%d') + datetime.timedelta(days=1) - datetime.timedelta(microseconds=1)
                    end_date = timezone.make_aware(naive_end, current_tz)
                except ValueError:
                    start_date = now - datetime.timedelta(days=30)
                    end_date = now
            else:
                start_date = now - datetime.timedelta(days=30)
                end_date = now
            
            context['start_date'] = start_date_str or start_date.date().isoformat()
            context['end_date'] = end_date_str or end_date.date().isoformat()

            # 3. Calculate Availability
            total = HistorialDisponibilidad.objects.filter(
                equipo=equipo,
                timestamp__range=(start_date, end_date)
            ).count()
            
            online = HistorialDisponibilidad.objects.filter(
                equipo=equipo,
                timestamp__range=(start_date, end_date),
                estado='ONLINE'
            ).count()
            
            availability = round((online / total * 100), 2) if total > 0 else 0
            context['availability'] = availability
            context['total_checks'] = total
            context['downtime_count'] = total - online

            # 4. Downtime Incidents - Group consecutive OFFLINE events
            all_history = HistorialDisponibilidad.objects.filter(
                equipo=equipo,
                timestamp__range=(start_date, end_date)
            ).order_by('timestamp').values('timestamp', 'estado')
            
            # Group consecutive OFFLINE periods into incidents
            incidents = []
            current_incident = None
            
            for record in all_history:
                if record['estado'] == 'OFFLINE':
                    if current_incident is None:
                        # Start new incident
                        current_incident = {
                            'start': record['timestamp'],
                            'end': record['timestamp']
                        }
                    else:
                        # Extend current incident
                        current_incident['end'] = record['timestamp']
                else:  # ONLINE
                    if current_incident is not None:
                        # Close incident and calculate duration
                        duration = current_incident['end'] - current_incident['start']
                        incidents.append({
                            'start': current_incident['start'],
                            'end': current_incident['end'],
                            'duration': duration,
                            'duration_str': str(duration).split('.')[0]  # Remove microseconds
                        })
                        current_incident = None
            
            # Close any remaining incident
            if current_incident is not None:
                duration = current_incident['end'] - current_incident['start']
                incidents.append({
                    'start': current_incident['start'],
                    'end': current_incident['end'],
                    'duration': duration,
                    'duration_str': str(duration).split('.')[0]
                })
            
            context['downtime_logs'] = incidents

            # 5. Chart Data
            chart_history = HistorialDisponibilidad.objects.filter(
                equipo=equipo,
                timestamp__range=(start_date, end_date)
            ).order_by('timestamp').values('timestamp', 'estado')
            
            chart_labels = []
            chart_values = []
            
            current_tz = timezone.get_current_timezone()
            
            for h in chart_history:
                local_dt = h['timestamp'].astimezone(current_tz)
                chart_labels.append(local_dt.strftime('%d/%m %H:%M'))
                chart_values.append(1 if h['estado'] == 'ONLINE' else 0)
            
            context['chart_labels'] = chart_labels
            context['chart_values'] = chart_values
            
        return context

        # 4. Total Devices Checked
        total_monitored = Equipo.objects.filter(estado='ACTIVO').count()

        context['total_outages_24h'] = total_outages_24h
        context['avg_latency_24h'] = round(avg_latency_24h, 2) if avg_latency_24h else 0
        context['worst_devices'] = worst_devices
        context['total_monitored'] = total_monitored
        
        return context

@admin_required_method
class ImportEquiposView(View):
    """View for importing equipment from XLSX files."""
    
    def get(self, request):
        """Display the import form."""
        form = EquipoImportForm()
        return render(request, 'monitor/import_equipos.html', {'form': form})
    
    def post(self, request):
        """Process the uploaded XLSX file."""
        form = EquipoImportForm(request.POST, request.FILES)
        
        if not form.is_valid():
            return render(request, 'monitor/import_equipos.html', {'form': form})
        
        archivo = request.FILES['archivo_xlsx']
        
        # Import logic
        from openpyxl import load_workbook
        from django.db import transaction
        import tempfile
        import os
        
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            for chunk in archivo.chunks():
                tmp_file.write(chunk)
            tmp_file_path = tmp_file.name
        
        try:
            # Load workbook with openpyxl (handles UTF-8 automatically)
            workbook = load_workbook(tmp_file_path, data_only=True)
            sheet = workbook.active
            
            # Parse headers (first row)
            headers_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            if not headers_row:
                return render(request, 'monitor/import_equipos.html', {
                    'form': form,
                    'error': 'El archivo está vacío o no tiene encabezados.'
                })
            
            # Normalize headers (lowercase, strip whitespace)
            headers = {self._normalize_header(h): idx for idx, h in enumerate(headers_row) if h}
            
            # Track results for preview
            duplicates = []
            new_records = []
            errors = []
            
            # Process data rows (skip header)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip completely empty rows
                    if all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    
                    # Extract data from row
                    equipo_data = self._extract_row_data(headers, row, row_idx)
                    
                    # Validate required fields
                    if not equipo_data.get('id_equipo'):
                        errors.append({
                            'row': row_idx,
                            'error': 'Campo requerido "ID Equipo" está vacío.'
                        })
                        continue
                    
                    if not equipo_data.get('ip'):
                        errors.append({
                            'row': row_idx,
                            'error': 'Campo requerido "IP" está vacío.'
                        })
                        continue
                    
                    # Check for duplicates
                    if Equipo.objects.filter(id_equipo=equipo_data['id_equipo']).exists():
                        errors.append({
                            'row': row_idx,
                            'error': f'Ya existe un equipo con ID "{equipo_data["id_equipo"]}".'
                        })
                        continue
                    
                    if Equipo.objects.filter(ip=equipo_data['ip']).exists():
                        errors.append({
                            'row': row_idx,
                            'error': f'Ya existe un equipo con IP "{equipo_data["ip"]}".'
                        })
                        continue
                    
                    # Get or create foreign keys
                    marca = None
                    if equipo_data.get('marca'):
                        marca, _ = Marca.objects.get_or_create(nombre=equipo_data['marca'])
                    
                    tipo = None
                    if equipo_data.get('tipo'):
                        tipo, _ = TipoEquipo.objects.get_or_create(nombre=equipo_data['tipo'])
                    
                    # Create equipment
                    with transaction.atomic():
                        equipo = Equipo.objects.create(
                            id_equipo=equipo_data['id_equipo'],
                            ip=equipo_data['ip'],
                            marca=marca,
                            tipo=tipo,
                            estado=equipo_data.get('estado', 'ACTIVO'),
                            medio_comunicacion=equipo_data.get('medio_comunicacion', 'FIBRA'),
                            latitud=equipo_data.get('latitud'),
                            longitud=equipo_data.get('longitud'),
                            direccion=equipo_data.get('direccion'),
                            poste=equipo_data.get('poste'),
                            piloto=equipo_data.get('piloto'),
                            canasta=equipo_data.get('canasta', False),
                            permisos=equipo_data.get('permisos', False),
                        )
                        created_equipos.append(equipo)
                        success_count += 1
                
                except Exception as e:
                    errors.append({
                        'row': row_idx,
                        'error': f'Error al procesar fila: {str(e)}'
                    })
            
            workbook.close()
            
            # Check if this is confirmation step
            if request.POST.get('confirm_import') == 'true':
                duplicate_action = request.POST.get('duplicate_action', 'skip')
                created, updated, skipped, import_errors = self._execute_import_with_action(
                    duplicates, new_records, errors, duplicate_action
                )
                
                # Clear session and temp file
                if 'import_temp_file' in request.session:
                    temp_file = request.session['import_temp_file']
                    if os.path.exists(temp_file):
                        os.unlink(temp_file)
                    del request.session['import_temp_file']
                
                # Show results
                return render(request, 'monitor/import_equipos.html', {
                    'form': EquipoImportForm(),
                    'created_count': created,
                    'updated_count': updated,
                    'skipped_count': skipped,
                    'error_count': len(import_errors),
                    'errors': import_errors[:50],
                    'total_errors': len(import_errors),
                    'import_completed': True,
                })
            
            # Store temp file path in session for confirmation
            request.session['import_temp_file'] = tmp_file_path
            
            # If no duplicates found, execute immediately
            if not duplicates:
                duplicate_action = 'skip'  # No duplicates to handle
                created, updated, skipped, import_errors = self._execute_import_with_action(
                    [], new_records, errors, duplicate_action
                )
                
                # Clean up
                if os.path.exists(tmp_file_path):
                    os.unlink(tmp_file_path)
                
                return render(request, 'monitor/import_equipos.html', {
                    'form': EquipoImportForm(),
                    'created_count': created,
                    'error_count': len(import_errors),
                    'errors': import_errors[:50],
                    'total_errors': len(import_errors),
                    'import_completed': True,
                })
            
            # Show preview with duplicates
            return render(request, 'monitor/import_equipos.html', {
                'form': form,
                'show_preview': True,
                'duplicate_count': len(duplicates),
                'new_count': len(new_records),
                'error_count': len(errors),
                'duplicates': duplicates[:50],
                'total_duplicates': len(duplicates),
                'validation_errors': errors[:20],
                # Store data for step 2
                'duplicates_data': duplicates,
                'new_records_data': new_records,
            })
        
        except Exception as e:
            return render(request, 'monitor/import_equipos.html', {
                'form': form,
                'error': f'Error al procesar el archivo: {str(e)}'
            })
        
        finally:
            # Clean up temp file
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)
    
    def _normalize_header(self, header):
        """Normalize header names for case-insensitive matching."""
        if not header:
            return ''
        # Remove accents and convert to lowercase
        normalized = str(header).lower().strip()
        # Common variations
        normalized = normalized.replace('ó', 'o').replace('í', 'i').replace('á', 'a')
        normalized = normalized.replace('é', 'e').replace('ú', 'u').replace('ñ', 'n')
        return normalized
    
    def _extract_row_data(self, headers, row, row_idx):
        """Extract and validate data from a row."""
        data = {}
        
        # Helper function to get cell value
        def get_value(header_variations):
            for variation in header_variations:
                if variation in headers:
                    idx = headers[variation]
                    if idx < len(row):
                        value = row[idx]
                        if value is not None:
                            return str(value).strip() if not isinstance(value, (int, float, bool)) else value
            return None
        
        # ID Equipo (required)
        data['id_equipo'] = get_value(['id equipo', 'id_equipo', 'idequipo', 'equipo'])
        
        # IP (required)
        data['ip'] = get_value(['ip', 'direccion ip', 'ip address'])
        
        # Marca (optional)
        data['marca'] = get_value(['marca', 'brand'])
        
        # Tipo (optional)
        data['tipo'] = get_value(['tipo', 'tipo equipo', 'type'])
        
        # Estado (optional)
        estado_val = get_value(['estado', 'status'])
        if estado_val:
            estado_upper = str(estado_val).upper()
            data['estado'] = 'ACTIVO' if estado_upper in ['ACTIVO', 'ACTIVE', '1', 'SI', 'SÍ'] else 'INACTIVO'
        
        # Medio comunicación (optional)
        medio_val = get_value(['medio comunicacion', 'medio_comunicacion', 'medio'])
        if medio_val:
            medio_upper = str(medio_val).upper()
            data['medio_comunicacion'] = 'FIBRA' if 'FIBRA' in medio_upper or 'FIBER' in medio_upper else 'CELULAR'
        
        # Coordinates (optional)
        lat_val = get_value(['latitud', 'lat', 'latitude'])
        if lat_val:
            try:
                data['latitud'] = float(lat_val)
            except (ValueError, TypeError):
                pass
        
        lon_val = get_value(['longitud', 'lon', 'lng', 'longitude'])
        if lon_val:
            try:
                data['longitud'] = float(lon_val)
            except (ValueError, TypeError):
                pass
        
        # Dirección (optional)
        data['direccion'] = get_value(['direccion', 'dirección', 'address'])
        
        # Poste (optional)
        data['poste'] = get_value(['poste', 'pole'])
        
        # Piloto (optional)
        data['piloto'] = get_value(['piloto', 'pilot'])
        
        # Canasta (optional boolean)
        canasta_val = get_value(['canasta', 'basket'])
        if canasta_val is not None:
            data['canasta'] = self._parse_boolean(canasta_val)
        
        # Permisos (optional boolean)
        permisos_val = get_value(['permisos', 'permits', 'permissions'])
        if permisos_val is not None:
            data['permisos'] = self._parse_boolean(permisos_val)
        
        return data
    
    def _parse_boolean(self, value):
        """Parse boolean values from various formats."""
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value != 0
        if isinstance(value, str):
            value_lower = value.lower().strip()
            return value_lower in ['si', 'sí', 'yes', 'true', '1', 'verdadero']
        return False
    
    def _execute_import_with_action(self, duplicates, new_records, validation_errors, duplicate_action):
        """Execute import with specified duplicate action."""
        from django.db import transaction
        
        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = list(validation_errors)
        
        # Process new records
        for record in new_records:
            try:
                with transaction.atomic():
                    equipo_data = record['data']
                    
                    marca = None
                    if equipo_data.get('marca'):
                        marca, _ = Marca.objects.get_or_create(nombre=equipo_data['marca'])
                    
                    tipo = None
                    if equipo_data.get('tipo'):
                        tipo, _ = TipoEquipo.objects.get_or_create(nombre=equipo_data['tipo'])
                    
                    Equipo.objects.create(
                        id_equipo=equipo_data['id_equipo'],
                        ip=equipo_data['ip'],
                        marca=marca,
                        tipo=tipo,
                        estado=equipo_data.get('estado', 'ACTIVO'),
                        medio_comunicacion=equipo_data.get('medio_comunicacion', 'FIBRA'),
                        latitud=equipo_data.get('latitud'),
                        longitud=equipo_data.get('longitud'),
                        direccion=equipo_data.get('direccion'),
                        poste=equipo_data.get('poste'),
                        piloto=equipo_data.get('piloto'),
                        canasta=equipo_data.get('canasta', False),
                        permisos=equipo_data.get('permisos', False),
                    )
                    created_count += 1
            except Exception as e:
                errors.append({'row': record['row'], 'error': f'Error al crear: {str(e)}'})
        
        # Process duplicates
        for duplicate in duplicates:
            try:
                if duplicate_action == 'skip':
                    skipped_count += 1
                elif duplicate_action == 'update':
                    with transaction.atomic():
                        existing = Equipo.objects.get(id_equipo=duplicate['id_equipo'])
                        self._merge_equipment_data(existing, duplicate['data'])
                        updated_count += 1
            except Exception as e:
                errors.append({'row': duplicate['row'], 'error': f'Error al actualizar: {str(e)}'})
        
        return created_count, updated_count, skipped_count, errors
    
    def _merge_equipment_data(self, existing_equipo, import_data):
        """Merge import data preserving existing values when import is empty."""
        if import_data.get('marca'):
            marca, _ = Marca.objects.get_or_create(nombre=import_data['marca'])
            existing_equipo.marca = marca
        
        if import_data.get('tipo'):
            tipo, _ = TipoEquipo.objects.get_or_create(nombre=import_data['tipo'])
            existing_equipo.tipo = tipo
        
        if import_data.get('ip'):
            existing_equipo.ip = import_data['ip']
        
        if import_data.get('estado'):
            existing_equipo.estado = import_data['estado']
        
        if import_data.get('medio_comunicacion'):
            existing_equipo.medio_comunicacion = import_data['medio_comunicacion']
        
        if import_data.get('latitud') is not None:
            existing_equipo.latitud = import_data['latitud']
        
        if import_data.get('longitud') is not None:
            existing_equipo.longitud = import_data['longitud']
        
        if import_data.get('direccion'):
            existing_equipo.direccion = import_data['direccion']
        
        if import_data.get('poste'):
            existing_equipo.poste = import_data['poste']
        
        if import_data.get('piloto'):
            existing_equipo.piloto = import_data['piloto']
        
        if 'canasta' in import_data:
            existing_equipo.canasta = import_data['canasta']
        
        if 'permisos' in import_data:
            existing_equipo.permisos = import_data['permisos']
        
        existing_equipo.save()


class DownloadImportTemplateView(View):
    """View to download XLSX import template."""
    
    def get(self, request):
        """Generate and return template XLSX file."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        import io
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Equipos"
        
        # Headers
        headers = [
            'ID Equipo', 'IP', 'Marca', 'Tipo', 'Estado', 
            'Medio Comunicación', 'Latitud', 'Longitud', 
            'Dirección', 'Poste', 'Piloto', 'Canasta', 'Permisos'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Get real data from database
        marcas = list(Marca.objects.values_list('nombre', flat=True)[:3])
        tipos = list(TipoEquipo.objects.values_list('nombre', flat=True)[:3])
        
        # Fallback to defaults if no data exists
        if not marcas:
            marcas = ['Honeywell', 'Itron', 'Trilliant']
        if not tipos:
            tipos = ['Colector', 'Medidor', 'Router']
        
        # Example data rows with real values
        examples = [
            {
                'ID Equipo': 'EQ-001',
                'IP': '192.168.1.100',
                'Marca': marcas[0] if len(marcas) > 0 else 'Honeywell',
                'Tipo': tipos[0] if len(tipos) > 0 else 'Colector',
                'Estado': 'ACTIVO',
                'Medio Comunicación': 'FIBRA',
                'Latitud': -0.2299,
                'Longitud': -78.5249,
                'Dirección': 'Av. República del Salvador N34-183',
                'Poste': 'P-001',
                'Piloto': 'José García',
                'Canasta': 'Sí',
                'Permisos': 'Sí'
            },
            {
                'ID Equipo': 'EQ-002',
                'IP': '192.168.1.101',
                'Marca': marcas[1] if len(marcas) > 1 else 'Itron',
                'Tipo': tipos[1] if len(tipos) > 1 else 'Medidor',
                'Estado': 'ACTIVO',
                'Medio Comunicación': 'CELULAR',
                'Latitud': -0.1807,
                'Longitud': -78.4678,
                'Dirección': 'Calle España y 10 de Agosto',
                'Poste': 'P-002',
                'Piloto': 'María López',
                'Canasta': 'No',
                'Permisos': 'Sí'
            }
        ]
        
        for row_num, example in enumerate(examples, 2):
            for col_num, header in enumerate(headers, 1):
                value = example.get(header, '')
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_importacion_equipos.xlsx"'
        
        return response


class PingDeviceView(View):
    def post(self, request, pk, *args, **kwargs):
        from .tasks import ping_host
        import json
        
        device = get_object_or_404(Equipo, pk=pk)
        
        # Run synchronous ping (fast timeout for UI)
        latency = ping_host(device.ip, timeout=2)
        
        status = 'ONLINE' if latency is not None else 'OFFLINE'
        
        # Update DB
        if status == 'ONLINE':
            device.last_seen = timezone.now()
            device.is_online = True
            # Update latency history if needed, or just update status
            HistorialDisponibilidad.objects.create(
                equipo=device,
                latencia_ms=latency,
                estado='ONLINE',
                packet_loss=0.0
            )
        else:
            device.is_online = False
            HistorialDisponibilidad.objects.create(
                equipo=device,
                latencia_ms=None,
                estado='OFFLINE',
                packet_loss=100.0
            )
        device.save()
        
        # Trigger client-side event for Toast
        response = HttpResponse(status=200)
        
        payload = {
            "showMessage": {
                "level": "success" if status == 'ONLINE' else "error",
                "message": f"Ping a {device.ip}: {status} ({latency}ms)" if latency else f"Ping a {device.ip}: {status}"
            }
        }
        response['HX-Trigger'] = json.dumps(payload)
        return response

class PingModalView(View):
    def get(self, request, pk, *args, **kwargs):
        device = get_object_or_404(Equipo, pk=pk)
        return render(request, 'monitor/partials/ping_modal.html', {'equipo': device})

class PingToolView(View):
    def get(self, request, pk, *args, **kwargs):
        from .tasks import ping_host
        import datetime
        from django.http import HttpResponse
        
        device = get_object_or_404(Equipo, pk=pk)
        latency = ping_host(device.ip, timeout=0.8) # Fast timeout for real-time feel
        
        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
        
        if latency is not None:
             html = f'<div class="mb-1"><span class="text-white-50">[{timestamp}]</span> Respuesta desde <b class="text-white">{device.ip}</b>: tiempo={latency}ms TTL=64</div>'
        else:
             html = f'<div class="mb-1"><span class="text-white-50">[{timestamp}]</span> <span class="text-danger">Tiempo de espera agotado para esta solicitud.</span></div>'
             
        return HttpResponse(html)


# User Management Views

@admin_required_method
class UsuarioListView(ListView):
    """View to list all users with their profiles."""
    model = User
    template_name = 'monitor/usuario_list.html'
    context_object_name = 'usuarios'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = User.objects.select_related('profile').filter(is_active=True)
        
        # Search functionality
        search_query = self.request.GET.get('q', '')
        if search_query:
            queryset = queryset.filter(
                Q(username__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(email__icontains=search_query)
            )
        
        # Filter by role
        role_filter = self.request.GET.get('role', '')
        if role_filter:
            queryset = queryset.filter(profile__role=role_filter)
        
        return queryset.order_by('username')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('q', '')
        context['role_filter'] = self.request.GET.get('role', '')
        return context


@admin_required_method
class UsuarioDetailView(DetailView):
    """View to display user profile details."""
    model = User
    template_name = 'monitor/usuario_detail.html'
    context_object_name = 'usuario'
    
    def get_queryset(self):
        return User.objects.select_related('profile')
    
    def get_object(self, queryset=None):
        user = super().get_object(queryset)
        # Ensure profile exists for legacy users
        if not hasattr(user, 'profile'):
            UserProfile.objects.create(user=user)
        return user


@admin_required_method
class UsuarioCreateView(View):
    """View to create a new user with profile."""
    
    def get(self, request):
        form = UserProfileForm()
        return render(request, 'monitor/usuario_form.html', {
            'form': form,
            'title': 'Crear Usuario',
            'is_create': True
        })
    
    def post(self, request):
        form = UserProfileForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('usuario_list')
        return render(request, 'monitor/usuario_form.html', {
            'form': form,
            'title': 'Crear Usuario',
            'is_create': True
        })


@admin_required_method
class UsuarioUpdateView(View):
    """View to update user and profile."""
    
    def get(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        # Create profile if it doesn't exist (for legacy users)
        profile, created = UserProfile.objects.get_or_create(user=user)
        form = UserProfileForm(instance=profile)
        return render(request, 'monitor/usuario_form.html', {
            'form': form,
            'usuario': user,
            'title': 'Editar Usuario',
            'is_create': False
        })
    
    def post(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        # Create profile if it doesn't exist (for legacy users)
        profile, created = UserProfile.objects.get_or_create(user=user)
        form = UserProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            return redirect('usuario_detail', pk=pk)
        return render(request, 'monitor/usuario_form.html', {
            'form': form,
            'usuario': user,
            'title': 'Editar Usuario',
            'is_create': False
        })


# Authentication Views

class LoginView(View):
    """View to handle user login."""
    
    def get(self, request):
        # Redirect if already logged in
        if request.user.is_authenticated:
            return redirect('dashboard')
        
        return render(request, 'monitor/login.html', {
            'next': request.GET.get('next', '')
        })
    
    def post(self, request):
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember = request.POST.get('remember')
        next_url = request.POST.get('next', '/')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # Handle "remember me"
            if not remember:
                request.session.set_expiry(0)  # Session expires on browser close
            else:
                request.session.set_expiry(1209600)  # 2 weeks
            
            messages.success(request, f'Bienvenido, {user.first_name or user.username}!')
            
            # Redirect to next or dashboard
            if next_url and next_url != '/login/':
                return redirect(next_url)
            return redirect('dashboard')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
            return render(request, 'monitor/login.html', {
                'form': {'errors': True},
                'next': next_url
            })


class LogoutView(View):
    """View to handle user logout."""
    
    def get(self, request):
        logout(request)
        messages.info(request, 'Ha cerrado sesión exitosamente.')
        return redirect('login')
    
    def post(self, request):
        logout(request)
        messages.info(request, 'Ha cerrado sesión exitosamente.')
        return redirect('login')



# Configuration and Master Data Views

@admin_required_method  
class ConfiguracionView(View):
    """View to display and edit global system configuration."""
    
    def get(self, request):
        config = ConfiguracionGlobal.load()
        form = ConfiguracionGlobalForm(instance=config)
        return render(request, 'monitor/configuracion.html', {
            'form': form,
            'config': config
        })
    
    def post(self, request):
        config = ConfiguracionGlobal.load()
        form = ConfiguracionGlobalForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuración actualizada exitosamente.')
            return redirect('configuracion')
        return render(request, 'monitor/configuracion.html', {
            'form': form,
            'config': config
        })


# Marcas CRUD

@admin_required_method
class MarcaListView(ListView):
    """List all equipment brands with equipment count."""
    model = Marca
    template_name = 'monitor/marca_list.html'
    context_object_name = 'marcas'
    
    def get_queryset(self):
        queryset = Marca.objects.annotate(
            equipment_count=Count('equipo')
        ).order_by('nombre')
        
        search = self.request.GET.get('q', '')
        if search:
            queryset = queryset.filter(nombre__icontains=search)
        
        return queryset


@admin_required_method
class MarcaCreateView(View):
    """Create a new brand."""
    
    def get(self, request):
        form = MarcaForm()
        return render(request, 'monitor/marca_form.html', {
            'form': form,
            'title': 'Nueva Marca',
            'is_create': True
        })
    
    def post(self, request):
        form = MarcaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'Marca "{form.cleaned_data["nombre"]}" creada exitosamente.')
            return redirect('marca_list')
        return render(request, 'monitor/marca_form.html', {
            'form': form,
            'title': 'Nueva Marca',
            'is_create': True
        })


@admin_required_method
class MarcaUpdateView(View):
    """Update an existing brand."""
    
    def get(self, request, pk):
        marca = get_object_or_404(Marca, pk=pk)
        form = MarcaForm(instance=marca)
        return render(request, 'monitor/marca_form.html', {
            'form': form,
            'marca': marca,
            'title': 'Editar Marca',
            'is_create': False
        })
    
    def post(self, request, pk):
        marca = get_object_or_404(Marca, pk=pk)
        form = MarcaForm(request.POST, instance=marca)
        if form.is_valid():
            form.save()
            messages.success(request, f'Marca "{marca.nombre}" actualizada exitosamente.')
            return redirect('marca_list')
        return render(request, 'monitor/marca_form.html', {
            'form': form,
            'marca': marca,
            'title': 'Editar Marca',
            'is_create': False
        })


@admin_required_method
class MarcaDeleteView(View):
    """Delete a brand (with protection if equipment exists)."""
    
    def get(self, request, pk):
        marca = get_object_or_404(Marca, pk=pk)
        equipment_count = marca.equipo_set.count()
        return render(request, 'monitor/marca_confirm_delete.html', {
            'marca': marca,
            'equipment_count': equipment_count
        })
    
    def post(self, request, pk):
        marca = get_object_or_404(Marca, pk=pk)
        equipment_count = marca.equipo_set.count()
        
        if equipment_count > 0:
            messages.error(request, f'No se puede eliminar la marca "{marca.nombre}" porque tiene {equipment_count} equipo(s) asignado(s).')
            return redirect('marca_list')
        
        nombre = marca.nombre
        marca.delete()
        messages.success(request, f'Marca "{nombre}" eliminada exitosamente.')
        return redirect('marca_list')


# TipoEquipo CRUD

@admin_required_method
class TipoEquipoListView(ListView):
    """List all equipment types with equipment count."""
    model = TipoEquipo
    template_name = 'monitor/tipoequipo_list.html'
    context_object_name = 'tipos'
    
    def get_queryset(self):
        queryset = TipoEquipo.objects.annotate(
            equipment_count=Count('equipo')
        ).order_by('nombre')
        
        search = self.request.GET.get('q', '')
        if search:
            queryset = queryset.filter(nombre__icontains=search)
        
        return queryset


@admin_required_method
class TipoEquipoCreateView(View):
    """Create a new equipment type."""
    
    def get(self, request):
        form = TipoEquipoForm()
        return render(request, 'monitor/tipoequipo_form.html', {
            'form': form,
            'title': 'Nuevo Tipo de Equipo',
            'is_create': True
        })
    
    def post(self, request):
        form = TipoEquipoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'Tipo "{form.cleaned_data["nombre"]}" creado exitosamente.')
            return redirect('tipo_list')
        return render(request, 'monitor/tipoequipo_form.html', {
            'form': form,
            'title': 'Nuevo Tipo de Equipo',
            'is_create': True
        })


@admin_required_method
class TipoEquipoUpdateView(View):
    """Update an existing equipment type."""
    
    def get(self, request, pk):
        tipo = get_object_or_404(TipoEquipo, pk=pk)
        form = TipoEquipoForm(instance=tipo)
        return render(request, 'monitor/tipoequipo_form.html', {
            'form': form,
            'tipo': tipo,
            'title': 'Editar Tipo de Equipo',
            'is_create': False
        })
    
    def post(self, request, pk):
        tipo = get_object_or_404(TipoEquipo, pk=pk)
        form = TipoEquipoForm(request.POST, instance=tipo)
        if form.is_valid():
            form.save()
            messages.success(request, f'Tipo "{tipo.nombre}" actualizado exitosamente.')
            return redirect('tipo_list')
        return render(request, 'monitor/tipoequipo_form.html', {
            'form': form,
            'tipo': tipo,
            'title': 'Editar Tipo de Equipo',
            'is_create': False
        })


@admin_required_method
class TipoEquipoDeleteView(View):
    """Delete an equipment type (with protection if equipment exists)."""
    
    def get(self, request, pk):
        tipo = get_object_or_404(TipoEquipo, pk=pk)
        equipment_count = tipo.equipo_set.count()
        return render(request, 'monitor/tipoequipo_confirm_delete.html', {
            'tipo': tipo,
            'equipment_count': equipment_count
        })
    
    def post(self, request, pk):
        tipo = get_object_or_404(TipoEquipo, pk=pk)
        equipment_count = tipo.equipo_set.count()
        
        if equipment_count > 0:
            messages.error(request, f'No se puede eliminar el tipo "{tipo.nombre}" porque tiene {equipment_count} equipo(s) asignado(s).')
            return redirect('tipo_list')
        
        nombre = tipo.nombre
        tipo.delete()
        messages.success(request, f'Tipo "{nombre}" eliminado exitosamente.')
        return redirect('tipo_list')


# Configuration and Master Data Views

@admin_required_method  
class ConfiguracionView(View):
    """View to display and edit global system configuration."""
    
    def get(self, request):
        config = ConfiguracionGlobal.load()
        form = ConfiguracionGlobalForm(instance=config)
        return render(request, 'monitor/configuracion.html', {
            'form': form,
            'config': config
        })
    
    def post(self, request):
        config = ConfiguracionGlobal.load()
        form = ConfiguracionGlobalForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuración actualizada exitosamente.')
            return redirect('configuracion')
        return render(request, 'monitor/configuracion.html', {
            'form': form,
            'config': config
        })


# Marcas CRUD

@admin_required_method
class MarcaListView(ListView):
    """List all equipment brands with equipment count."""
    model = Marca
    template_name = 'monitor/marca_list.html'
    context_object_name = 'marcas'
    
    def get_queryset(self):
        queryset = Marca.objects.annotate(
            equipment_count=Count('equipo')
        ).order_by('nombre')
        
        search = self.request.GET.get('q', '')
        if search:
            queryset = queryset.filter(nombre__icontains=search)
        
        return queryset


@admin_required_method
class MarcaCreateView(View):
    """Create a new brand."""
    
    def get(self, request):
        form = MarcaForm()
        return render(request, 'monitor/marca_form.html', {
            'form': form,
            'title': 'Nueva Marca',
            'is_create': True
        })
    
    def post(self, request):
        form = MarcaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'Marca "{form.cleaned_data["nombre"]}" creada exitosamente.')
            return redirect('marca_list')
        return render(request, 'monitor/marca_form.html', {
            'form': form,
            'title': 'Nueva Marca',
            'is_create': True
        })


@admin_required_method
class MarcaUpdateView(View):
    """Update an existing brand."""
    
    def get(self, request, pk):
        marca = get_object_or_404(Marca, pk=pk)
        form = MarcaForm(instance=marca)
        return render(request, 'monitor/marca_form.html', {
            'form': form,
            'marca': marca,
            'title': 'Editar Marca',
            'is_create': False
        })
    
    def post(self, request, pk):
        marca = get_object_or_404(Marca, pk=pk)
        form = MarcaForm(request.POST, instance=marca)
        if form.is_valid():
            form.save()
            messages.success(request, f'Marca "{marca.nombre}" actualizada exitosamente.')
            return redirect('marca_list')
        return render(request, 'monitor/marca_form.html', {
            'form': form,
            'marca': marca,
            'title': 'Editar Marca',
            'is_create': False
        })


@admin_required_method
class MarcaDeleteView(View):
    """Delete a brand (with protection if equipment exists)."""
    
    def get(self, request, pk):
        marca = get_object_or_404(Marca, pk=pk)
        equipment_count = marca.equipo_set.count()
        return render(request, 'monitor/marca_confirm_delete.html', {
            'marca': marca,
            'equipment_count': equipment_count
        })
    
    def post(self, request, pk):
        marca = get_object_or_404(Marca, pk=pk)
        equipment_count = marca.equipo_set.count()
        
        if equipment_count > 0:
            messages.error(request, f'No se puede eliminar la marca "{marca.nombre}" porque tiene {equipment_count} equipo(s) asignado(s).')
            return redirect('marca_list')
        
        nombre = marca.nombre
        marca.delete()
        messages.success(request, f'Marca "{nombre}" eliminada exitosamente.')
        return redirect('marca_list')


# TipoEquipo CRUD

@admin_required_method
class TipoEquipoListView(ListView):
    """List all equipment types with equipment count."""
    model = TipoEquipo
    template_name = 'monitor/tipoequipo_list.html'
    context_object_name = 'tipos'
    
    def get_queryset(self):
        queryset = TipoEquipo.objects.annotate(
            equipment_count=Count('equipo')
        ).order_by('nombre')
        
        search = self.request.GET.get('q', '')
        if search:
            queryset = queryset.filter(nombre__icontains=search)
        
        return queryset


@admin_required_method
class TipoEquipoCreateView(View):
    """Create a new equipment type."""
    
    def get(self, request):
        form = TipoEquipoForm()
        return render(request, 'monitor/tipoequipo_form.html', {
            'form': form,
            'title': 'Nuevo Tipo de Equipo',
            'is_create': True
        })
    
    def post(self, request):
        form = TipoEquipoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'Tipo "{form.cleaned_data["nombre"]}" creado exitosamente.')
            return redirect('tipo_list')
        return render(request, 'monitor/tipoequipo_form.html', {
            'form': form,
            'title': 'Nuevo Tipo de Equipo',
            'is_create': True
        })


@admin_required_method
class TipoEquipoUpdateView(View):
    """Update an existing equipment type."""
    
    def get(self, request, pk):
        tipo = get_object_or_404(TipoEquipo, pk=pk)
        form = TipoEquipoForm(instance=tipo)
        return render(request, 'monitor/tipoequipo_form.html', {
            'form': form,
            'tipo': tipo,
            'title': 'Editar Tipo de Equipo',
            'is_create': False
        })
    
    def post(self, request, pk):
        tipo = get_object_or_404(TipoEquipo, pk=pk)
        form = TipoEquipoForm(request.POST, instance=tipo)
        if form.is_valid():
            form.save()
            messages.success(request, f'Tipo "{tipo.nombre}" actualizado exitosamente.')
            return redirect('tipo_list')
        return render(request, 'monitor/tipoequipo_form.html', {
            'form': form,
            'tipo': tipo,
            'title': 'Editar Tipo de Equipo',
            'is_create': False
        })


@admin_required_method
class TipoEquipoDeleteView(View):
    """Delete an equipment type (with protection if equipment exists)."""
    
    def get(self, request, pk):
        tipo = get_object_or_404(TipoEquipo, pk=pk)
        equipment_count = tipo.equipo_set.count()
        return render(request, 'monitor/tipoequipo_confirm_delete.html', {
            'tipo': tipo,
            'equipment_count': equipment_count
        })
    
    def post(self, request, pk):
        tipo = get_object_or_404(TipoEquipo, pk=pk)
        equipment_count = tipo.equipo_set.count()
        
        if equipment_count > 0:
            messages.error(request, f'No se puede eliminar el tipo "{tipo.nombre}" porque tiene {equipment_count} equipo(s) asignado(s).')
            return redirect('tipo_list')
        
        nombre = tipo.nombre
        tipo.delete()
        messages.success(request, f'Tipo "{nombre}" eliminado exitosamente.')
        return redirect('tipo_list')


# Equipment CRUD Views

@admin_required_method
class EquipoCreateView(View):
    """Create a new equipment."""
    
    def get(self, request):
        form = EquipoForm()
        return render(request, 'monitor/equipo_form.html', {
            'form': form,
            'title': 'Nuevo Equipo',
            'is_create': True
        })
    
    def post(self, request):
        form = EquipoForm(request.POST)
        if form.is_valid():
            equipo = form.save()
            messages.success(request, f'Equipo "{equipo.id_equipo}" creado exitosamente.')
            return redirect('equipo_detail', pk=equipo.pk)
        return render(request, 'monitor/equipo_form.html', {
            'form': form,
            'title': 'Nuevo Equipo',
            'is_create': True
        })


@admin_required_method
class EquipoUpdateView(View):
    """Update an existing equipment."""
    
    def get(self, request, pk):
        equipo = get_object_or_404(Equipo, pk=pk)
        form = EquipoForm(instance=equipo)
        return render(request, 'monitor/equipo_form.html', {
            'form': form,
            'equipo': equipo,
            'title': 'Editar Equipo',
            'is_create': False
        })
    
    def post(self, request, pk):
        equipo = get_object_or_404(Equipo, pk=pk)
        form = EquipoForm(request.POST, instance=equipo)
        if form.is_valid():
            form.save()
            messages.success(request, f'Equipo "{equipo.id_equipo}" actualizado exitosamente.')
            return redirect('equipo_detail', pk=equipo.pk)
        return render(request, 'monitor/equipo_form.html', {
            'form': form,
            'equipo': equipo,
            'title': 'Editar Equipo',
            'is_create': False
        })


@admin_required_method
class EquipoDeleteView(View):
    """Delete an equipment with confirmation."""
    
    def get(self, request, pk):
        equipo = get_object_or_404(Equipo, pk=pk)
        history_count = equipo.historial.count()
        return render(request, 'monitor/equipo_confirm_delete.html', {
            'equipo': equipo,
            'history_count': history_count
        })
    
    def post(self, request, pk):
        equipo = get_object_or_404(Equipo, pk=pk)
        id_equipo = equipo.id_equipo
        equipo.delete()  # Cascades to historical data
        messages.success(request, f'Equipo "{id_equipo}" eliminado exitosamente.')
        return redirect('equipo_list')


# Password Change View

@login_required_method
class ChangePasswordView(View):
    """Allow authenticated users to change their password."""
    
    def get(self, request):
        form = PasswordChangeForm(user=request.user)
        return render(request, 'monitor/change_password.html', {
            'form': form
        })
    
    def post(self, request):
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            form.save()
            # Update session hash to prevent logout
            update_session_auth_hash(request, request.user)
            messages.success(request, 'Tu contraseña ha sido actualizada exitosamente.')
            return redirect('dashboard')
        return render(request, 'monitor/change_password.html', {
            'form': form
        })


# User Profile Management View

@login_required_method
class MyProfileView(View):
    """Allow users to view and edit their own profile."""
    
    def get(self, request):
        profile = request.user.profile
        form = UserProfileForm(instance=profile)
        return render(request, 'monitor/my_profile.html', {
            'form': form,
            'profile': profile
        })
    
    def post(self, request):
        profile = request.user.profile
        form = UserProfileForm(
            request.POST,
            request.FILES,
            instance=profile
        )
        if form.is_valid():
            form.save()
            messages.success(request, 'Tu perfil ha sido actualizado exitosamente.')
            return redirect('my_profile')
        return render(request, 'monitor/my_profile.html', {
            'form': form,
            'profile': profile
        })


# ==================== BILLING CALENDAR VIEWS ====================

import calendar
from datetime import date
from .models import Porcion, EventoFacturacion, CicloFacturacion
from .forms import PorcionForm, EventoFacturacionForm

@login_required_method
class CalendarioView(TemplateView):
    """View to display the monthly billing calendar."""
    template_name = 'monitor/calendario.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get year and month from URL or use current
        anio = self.kwargs.get('anio') or date.today().year
        mes = self.kwargs.get('mes') or date.today().month
        
        # Ensure valid month
        mes = max(1, min(12, mes))
        
        context['anio'] = anio
        context['mes'] = mes
        
        # Month names in Spanish
        meses_es = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
            5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
            9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }
        context['mes_nombre'] = meses_es[mes]
        
        # Calculate previous/next month
        if mes == 1:
            context['prev_mes'] = 12
            context['prev_anio'] = anio - 1
        else:
            context['prev_mes'] = mes - 1
            context['prev_anio'] = anio
            
        if mes == 12:
            context['next_mes'] = 1
            context['next_anio'] = anio + 1
        else:
            context['next_mes'] = mes + 1
            context['next_anio'] = anio
        
        # Get calendar data
        cal = calendar.monthcalendar(anio, mes)
        context['calendar'] = cal
        
        # Get all events for this month
        eventos = EventoFacturacion.objects.filter(
            fecha__year=anio,
            fecha__month=mes
        ).select_related('porcion', 'ciclo')
        
        # Group events by day
        eventos_por_dia = {}
        for evento in eventos:
            dia = evento.fecha.day
            if dia not in eventos_por_dia:
                eventos_por_dia[dia] = []
            eventos_por_dia[dia].append({
                'id': evento.id,
                'nombre': evento.get_display_name(),
                'color': evento.get_color(),
                'tipo': evento.tipo_evento,
                'porcion': evento.porcion.nombre
            })
        
        context['eventos_por_dia'] = eventos_por_dia
        context['today'] = date.today()
        
        return context


@login_required_method
class EventoListView(ListView):
    """View to list all billing events."""
    model = EventoFacturacion
    template_name = 'monitor/evento_list.html'
    context_object_name = 'eventos'
    paginate_by = 20
    
    def get_queryset(self):
        qs = super().get_queryset().select_related('porcion', 'ciclo').order_by('-fecha', '-created_at')
        
        # Filter by tipo_evento if specified
        tipo = self.request.GET.get('tipo')
        if tipo:
            qs = qs.filter(tipo_evento=tipo)
        
        # Filter by porcion if specified
        porcion_id = self.request.GET.get('porcion')
        if porcion_id:
            qs = qs.filter(porcion_id=porcion_id)
        
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tipo_filter'] = self.request.GET.get('tipo', '')
        context['porcion_filter'] = self.request.GET.get('porcion', '')
        context['porciones'] = Porcion.objects.all().order_by('nombre')
        return context



@admin_required_method
class EventoCreateView(View):
    """View to create billing events."""
    
    def get(self, request):
        form = EventoFacturacionForm()
        return render(request, 'monitor/evento_form.html', {'form': form})
    
    def post(self, request):
        form = EventoFacturacionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Evento de facturación creado exitosamente.')
            return redirect('calendario')
        return render(request, 'monitor/evento_form.html', {'form': form})


@admin_required_method
class EventoUpdateView(View):
    """View to update billing events."""
    
    def get(self, request, pk):
        evento = get_object_or_404(EventoFacturacion, pk=pk)
        form = EventoFacturacionForm(instance=evento)
        return render(request, 'monitor/evento_form.html', {
            'form': form,
            'evento': evento
        })
    
    def post(self, request, pk):
        evento = get_object_or_404(EventoFacturacion, pk=pk)
        form = EventoFacturacionForm(request.POST, instance=evento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Evento actualizado exitosamente.')
            return redirect('calendario')
        return render(request, 'monitor/evento_form.html', {
            'form': form,
            'evento': evento
        })


@admin_required_method
class EventoDeleteView(View):
    """View to delete billing events."""
    
    def post(self, request, pk):
        evento = get_object_or_404(EventoFacturacion, pk=pk)
        evento.delete()
        messages.success(request, 'Evento eliminado exitosamente.')
        return redirect('calendario')


# ==================== PORTION MANAGEMENT VIEWS ====================

@admin_required_method
class PorcionListView(ListView):
    """View to list all billing portions."""
    model = Porcion
    template_name = 'monitor/porcion_list.html'
    context_object_name = 'porciones'
    paginate_by = 20
    
    def get_queryset(self):
        qs = super().get_queryset()
        
        # Filter by type if specified
        tipo = self.request.GET.get('tipo')
        if tipo:
            qs = qs.filter(tipo=tipo)
        
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tipo_filter'] = self.request.GET.get('tipo', '')
        return context


@admin_required_method
class PorcionCreateView(View):
    """View to create portions."""
    
    def get(self, request):
        form = PorcionForm()
        return render(request, 'monitor/porcion_form.html', {'form': form})
    
    def post(self, request):
        form = PorcionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Porción creada exitosamente.')
            return redirect('porcion_list')
        return render(request, 'monitor/porcion_form.html', {'form': form})


@admin_required_method
class PorcionUpdateView(View):
    """View to update portions."""
    
    def get(self, request, pk):
        porcion = get_object_or_404(Porcion, pk=pk)
        form = PorcionForm(instance=porcion)
        return render(request, 'monitor/porcion_form.html', {
            'form': form,
            'porcion': porcion
        })
    
    def post(self, request, pk):
        porcion = get_object_or_404(Porcion, pk=pk)
        form = PorcionForm(request.POST, instance=porcion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Porción actualizada exitosamente.')
            return redirect('porcion_list')
        return render(request, 'monitor/porcion_form.html', {
            'form': form,
            'porcion': porcion
        })


@admin_required_method
class PorcionDeleteView(View):
    """View to delete portions."""
    
    def post(self, request, pk):
        porcion = get_object_or_404(Porcion, pk=pk)
        porcion.delete()
        messages.success(request, 'Porción eliminada exitosamente.')
        return redirect('porcion_list')

