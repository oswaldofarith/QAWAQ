from django.views import View
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.contrib import messages
import os
import tempfile
import pandas as pd
import openpyxl
import re
import logging
import traceback
from django.db import transaction
from ..forms import EquipoImportForm
from ..models import Equipo, Marca, TipoEquipo, Medidor, Porcion
from ..decorators import admin_required_method

@admin_required_method
class ImportEquiposView(View):
    """View for importing equipment from XLSX files."""
    
    def get(self, request):
        """Display the import form."""
        form = EquipoImportForm()
        return render(request, 'monitor/import_equipos.html', {'form': form})
    
    def post(self, request):
        """Process the uploaded XLSX file."""
        form = EquipoImportForm(request.POST, request.FILES)
        
        if not form.is_valid():
            return render(request, 'monitor/import_equipos.html', {'form': form})
        
        archivo = request.FILES['archivo_xlsx']
        
        # Import logic
        from openpyxl import load_workbook
        from django.db import transaction
        
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            for chunk in archivo.chunks():
                tmp_file.write(chunk)
            tmp_file_path = tmp_file.name
        
        try:
            # Load workbook with openpyxl (handles UTF-8 automatically)
            workbook = load_workbook(tmp_file_path, data_only=True)
            sheet = workbook.active
            
            # Parse headers (first row)
            headers_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            if not headers_row:
                return render(request, 'monitor/import_equipos.html', {
                    'form': form,
                    'error': 'El archivo está vacío o no tiene encabezados.'
                })
            
            # Normalize headers (lowercase, strip whitespace)
            headers = {self._normalize_header(h): idx for idx, h in enumerate(headers_row) if h}
            
            # Track results for preview
            duplicates = []
            new_records = []
            errors = []
            
            # Process data rows (skip header)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip completely empty rows
                    if all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    
                    # Extract data from row
                    equipo_data = self._extract_row_data(headers, row, row_idx)
                    
                    # Validate required fields
                    if not equipo_data.get('id_equipo'):
                        errors.append({
                            'row': row_idx,
                            'error': 'Campo requerido "ID Equipo" está vacío.'
                        })
                        continue
                    
                    if not equipo_data.get('ip'):
                        errors.append({
                            'row': row_idx,
                            'error': 'Campo requerido "IP" está vacío.'
                        })
                        continue
                    
                    # Check for duplicates
                    if Equipo.objects.filter(id_equipo=equipo_data['id_equipo']).exists():
                        duplicates.append(equipo_data)
                        continue
                    
                    if Equipo.objects.filter(ip=equipo_data['ip']).exists():
                        duplicates.append(equipo_data)
                        continue
                    
                    # Add to new records for processing
                    new_records.append(equipo_data)
                
                except Exception as e:
                    errors.append({
                        'row': row_idx,
                        'error': f'Error al procesar fila: {str(e)}'
                    })
            
            workbook.close()
            
            # Check if this is confirmation step
            if request.POST.get('confirm_import') == 'true':
                duplicate_action = request.POST.get('duplicate_action', 'skip')
                stats = self._execute_import_with_action(
                    duplicates, new_records, errors, duplicate_action
                )
                
                # Clear session and temp file
                if 'import_temp_file' in request.session:
                    temp_file = request.session['import_temp_file']
                    if os.path.exists(temp_file):
                        os.unlink(temp_file)
                    del request.session['import_temp_file']
                
                # Show summary page
                return render(request, 'monitor/equipment_import_summary.html', {
                    'stats': stats,
                    'success': True
                })
            
            # Store temp file path in session for confirmation
            request.session['import_temp_file'] = tmp_file_path
            
            # If no duplicates found, execute immediately
            if not duplicates:
                duplicate_action = 'skip'  # No duplicates to handle
                stats = self._execute_import_with_action(
                    [], new_records, errors, duplicate_action
                )
                
                # Clean up
                if os.path.exists(tmp_file_path):
                    os.unlink(tmp_file_path)
                
                # Show summary page
                return render(request, 'monitor/equipment_import_summary.html', {
                    'stats': stats,
                    'success': True
                })
            
            # Show preview with duplicates
            return render(request, 'monitor/import_equipos.html', {
                'form': form,
                'show_preview': True,
                'duplicate_count': len(duplicates),
                'new_count': len(new_records),
                'error_count': len(errors),
                'duplicates': duplicates[:50],
                'total_duplicates': len(duplicates),
                'validation_errors': errors[:20],
                # Store data for step 2
                'duplicates_data': duplicates,
                'new_records_data': new_records,
            })
        
        except Exception as e:
            return render(request, 'monitor/import_equipos.html', {
                'form': form,
                'error': f'Error al procesar el archivo: {str(e)}'
            })
        
        finally:
            # Clean up temp file
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)
    
    def _normalize_header(self, header):
        """Normalize header names for case-insensitive matching."""
        if not header:
            return ''
        # Remove accents and convert to lowercase
        normalized = str(header).lower().strip()
        # Common variations
        normalized = normalized.replace('ó', 'o').replace('í', 'i').replace('á', 'a')
        normalized = normalized.replace('é', 'e').replace('ú', 'u').replace('ñ', 'n')
        return normalized
    
    def _extract_row_data(self, headers, row, row_idx):
        """Extract and validate data from a row."""
        data = {}
        
        # Helper function to get cell value
        def get_value(header_variations):
            for variation in header_variations:
                if variation in headers:
                    idx = headers[variation]
                    if idx < len(row):
                        value = row[idx]
                        if value is not None:
                            return str(value).strip() if not isinstance(value, (int, float, bool)) else value
            return None
        
        # ID Equipo (required)
        data['id_equipo'] = get_value(['id equipo', 'id_equipo', 'idequipo', 'equipo'])
        
        # IP (required)
        data['ip'] = get_value(['ip', 'direccion ip', 'ip address'])
        
        # Marca (optional)
        data['marca'] = get_value(['marca', 'brand'])
        
        # Tipo (optional)
        data['tipo'] = get_value(['tipo', 'tipo equipo', 'type'])
        
        # Estado (optional)
        estado_val = get_value(['estado', 'status'])
        if estado_val:
            estado_upper = str(estado_val).upper()
            data['estado'] = 'ACTIVO' if estado_upper in ['ACTIVO', 'ACTIVE', '1', 'SI', 'SÍ'] else 'INACTIVO'
        
        # Medio comunicación (optional)
        medio_val = get_value(['medio comunicacion', 'medio_comunicacion', 'medio'])
        if medio_val:
            medio_upper = str(medio_val).upper()
            data['medio_comunicacion'] = 'FIBRA' if 'FIBRA' in medio_upper or 'FIBER' in medio_upper else 'CELULAR'
        
        # Coordinates (optional)
        lat_val = get_value(['latitud', 'lat', 'latitude'])
        if lat_val:
            try:
                data['latitud'] = float(lat_val)
            except (ValueError, TypeError):
                pass
        
        lon_val = get_value(['longitud', 'lon', 'lng', 'longitude'])
        if lon_val:
            try:
                data['longitud'] = float(lon_val)
            except (ValueError, TypeError):
                pass
        
        # Dirección (optional)
        data['direccion'] = get_value(['direccion', 'dirección', 'address'])
        
        # Poste (optional)
        data['poste'] = get_value(['poste', 'pole'])
        
        # Piloto (optional)
        data['piloto'] = get_value(['piloto', 'pilot'])
        
        # Canasta (optional boolean)
        canasta_val = get_value(['canasta', 'basket'])
        if canasta_val is not None:
            data['canasta'] = self._parse_boolean(canasta_val)
        
        # Permisos (optional boolean)
        permisos_val = get_value(['permisos', 'permits', 'permissions'])
        if permisos_val is not None:
            data['permisos'] = self._parse_boolean(permisos_val)
        
        return data
    
    def _parse_boolean(self, value):
        """Parse boolean values from various formats."""
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value != 0
        if isinstance(value, str):
            value_lower = value.lower().strip()
            return value_lower in ['si', 'sí', 'yes', 'true', '1', 'verdadero']
        return False
    
    def _execute_import_with_action(self, duplicates, new_records, validation_errors, duplicate_action):
        """Execute import with specified duplicate action."""
        from django.db import transaction
        
        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = list(validation_errors)
        
        # Track detailed lists
        created_equipos = []
        updated_equipos = []
        rejected_equipos = []
        
        # Add validation errors to rejected list
        for error in validation_errors:
            rejected_equipos.append({
                'row': error.get('row', 'N/A'),
                'id_equipo': error.get('id_equipo', 'N/A'),
                'reason': error.get('error', 'Error de validación')
            })
        
        # Process new records
        for record in new_records:
            try:
                with transaction.atomic():
                    equipo_data = record['data']
                    
                    marca = None
                    if equipo_data.get('marca'):
                        # Case-insensitive lookup
                        marca = Marca.objects.filter(nombre__iexact=equipo_data['marca']).first()
                        if not marca:
                            marca = Marca.objects.create(nombre=equipo_data['marca'])
                    
                    tipo = None
                    if equipo_data.get('tipo'):
                        # Case-insensitive lookup
                        tipo = TipoEquipo.objects.filter(nombre__iexact=equipo_data['tipo']).first()
                        if not tipo:
                            tipo = TipoEquipo.objects.create(nombre=equipo_data['tipo'])
                    
                    Equipo.objects.create(
                        id_equipo=equipo_data['id_equipo'],
                        ip=equipo_data['ip'],
                        marca=marca,
                        tipo=tipo,
                        estado=equipo_data.get('estado', 'ACTIVO'),
                        medio_comunicacion=equipo_data.get('medio_comunicacion', 'FIBRA'),
                        latitud=equipo_data.get('latitud'),
                        longitud=equipo_data.get('longitud'),
                        direccion=equipo_data.get('direccion'),
                        poste=equipo_data.get('poste'),
                        piloto=equipo_data.get('piloto'),
                        canasta=equipo_data.get('canasta', False),
                        permisos=equipo_data.get('permisos', False),
                    )
                    created_count += 1
                    created_equipos.append({
                        'id_equipo': equipo_data['id_equipo'],
                        'ip': equipo_data['ip']
                    })
            except Exception as e:
                errors.append({'row': record['row'], 'error': f'Error al crear: {str(e)}'})
                rejected_equipos.append({
                    'row': record.get('row', 'N/A'),
                    'id_equipo': equipo_data.get('id_equipo', 'N/A'),
                    'reason': f'Error al crear: {str(e)}'
                })
        
        # Process duplicates
        for duplicate in duplicates:
            try:
                if duplicate_action == 'skip':
                    skipped_count += 1
                    rejected_equipos.append({
                        'row': duplicate.get('row', 'N/A'),
                        'id_equipo': duplicate.get('id_equipo', 'N/A'),
                        'reason': 'Equipo duplicado (omitido por el usuario)'
                    })
                elif duplicate_action == 'update':
                    with transaction.atomic():
                        existing = Equipo.objects.get(id_equipo=duplicate['id_equipo'])
                        self._merge_equipment_data(existing, duplicate['data'])
                        updated_count += 1
                        updated_equipos.append({
                            'id_equipo': duplicate['id_equipo'],
                            'ip': existing.ip
                        })
            except Exception as e:
                errors.append({'row': duplicate['row'], 'error': f'Error al actualizar: {str(e)}'})
                rejected_equipos.append({
                    'row': duplicate.get('row', 'N/A'),
                    'id_equipo': duplicate.get('id_equipo', 'N/A'),
                    'reason': f'Error al actualizar: {str(e)}'
                })
        
        # Return comprehensive statistics
        return {
            'created': len(created_equipos),
            'updated': len(updated_equipos),
            'rejected': len(rejected_equipos),
            'created_equipos': created_equipos,
            'updated_equipos': updated_equipos,
            'rejected_equipos': rejected_equipos
        }
    
    def _merge_equipment_data(self, existing_equipo, import_data):
        """Merge import data preserving existing values when import is empty."""
        if import_data.get('marca'):
            marca, _ = Marca.objects.get_or_create(nombre=import_data['marca'])
            existing_equipo.marca = marca
        
        if import_data.get('tipo'):
            tipo, _ = TipoEquipo.objects.get_or_create(nombre=import_data['tipo'])
            existing_equipo.tipo = tipo
        
        if import_data.get('ip'):
            existing_equipo.ip = import_data['ip']
        
        if import_data.get('estado'):
            existing_equipo.estado = import_data['estado']
        
        if import_data.get('medio_comunicacion'):
            existing_equipo.medio_comunicacion = import_data['medio_comunicacion']
        
        if import_data.get('latitud') is not None:
            existing_equipo.latitud = import_data['latitud']
        
        if import_data.get('longitud') is not None:
            existing_equipo.longitud = import_data['longitud']
        
        if import_data.get('direccion'):
            existing_equipo.direccion = import_data['direccion']
        
        if import_data.get('poste'):
            existing_equipo.poste = import_data['poste']
        
        if import_data.get('piloto'):
            existing_equipo.piloto = import_data['piloto']
        
        if 'canasta' in import_data:
            existing_equipo.canasta = import_data['canasta']
        
        if 'permisos' in import_data:
            existing_equipo.permisos = import_data['permisos']
        
        existing_equipo.save()


class DownloadImportTemplateView(View):
    """View to download XLSX import template."""
    
    def get(self, request):
        """Generate and return template XLSX file."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        import io
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Equipos"
        
        # Headers
        headers = [
            'ID Equipo', 'IP', 'Marca', 'Tipo', 'Estado', 
            'Medio Comunicación', 'Latitud', 'Longitud', 
            'Dirección', 'Poste', 'Piloto', 'Canasta', 'Permisos'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Get real data from database
        marcas = list(Marca.objects.values_list('nombre', flat=True)[:3])
        tipos = list(TipoEquipo.objects.values_list('nombre', flat=True)[:3])
        
        # Fallback to defaults if no data exists
        if not marcas:
            marcas = ['Honeywell', 'Itron', 'Trilliant']
        if not tipos:
            tipos = ['Colector', 'Medidor', 'Router']
        
        # Example data rows with real values
        examples = [
            {
                'ID Equipo': 'EQ-001',
                'IP': '192.168.1.100',
                'Marca': marcas[0] if len(marcas) > 0 else 'Honeywell',
                'Tipo': tipos[0] if len(tipos) > 0 else 'Colector',
                'Estado': 'ACTIVO',
                'Medio Comunicación': 'FIBRA',
                'Latitud': -0.2299,
                'Longitud': -78.5249,
                'Dirección': 'Av. República del Salvador N34-183',
                'Poste': 'P-001',
                'Piloto': 'José García',
                'Canasta': 'Sí',
                'Permisos': 'Sí'
            },
            {
                'ID Equipo': 'EQ-002',
                'IP': '192.168.1.101',
                'Marca': marcas[1] if len(marcas) > 1 else 'Itron',
                'Tipo': tipos[1] if len(tipos) > 1 else 'Medidor',
                'Estado': 'ACTIVO',
                'Medio Comunicación': 'CELULAR',
                'Latitud': -0.1807,
                'Longitud': -78.4678,
                'Dirección': 'Calle España y 10 de Agosto',
                'Poste': 'P-002',
                'Piloto': 'María López',
                'Canasta': 'No',
                'Permisos': 'Sí'
            }
        ]
        
        for row_num, example in enumerate(examples, 2):
            for col_num, header in enumerate(headers, 1):
                value = example.get(header, '')
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_importacion_equipos.xlsx"'
        
        return response

@admin_required_method
class ImportMedidoresView(View):
    """View to import AMI meters from XLSX files with data transformations."""
    
    def get(self, request):
        return render(request, 'monitor/import_medidores.html', {
            'total_medidores': Medidor.objects.count()
        })
    
    def post(self, request):
        logger = logging.getLogger(__name__)
        
        if 'file' not in request.FILES:
            messages.error(request, 'No se ha seleccionado ningún archivo.')
            return redirect('import_medidores')
        
        xlsx_file = request.FILES['file']
        
        # Log file details
        logger.info(f"Processing file: {xlsx_file.name}, Size: {xlsx_file.size} bytes")
        
        # Validate file extension
        if not xlsx_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'El archivo debe ser formato XLSX o XLS.')
            return redirect('import_medidores')
        
        # Check file size (100 MB limit)
        max_size = 100 * 1024 * 1024  # 100 MB
        if xlsx_file.size > max_size:
            messages.error(request, f'El archivo es demasiado grande ({xlsx_file.size / 1024 / 1024:.2f} MB). Máximo permitido: 100 MB.')
            return redirect('import_medidores')
        
        try:
            logger.info("Starting XLSX processing...")
            # Process XLSX file
            processed_data = self._process_xlsx_data(xlsx_file)
            logger.info(f"Processed {len(processed_data)} records from XLSX")
            
            if not processed_data:
                messages.warning(request, 'No se encontraron datos válidos en el archivo.')
                return redirect('import_medidores')
            
            logger.info("Starting data import...")
            # Import data (delete existing and create new) - returns statistics dict
            stats = self._import_data(processed_data)
            logger.info(f"Import complete: {stats['imported']} imported, {stats['rejected_by_marca']['total']} rejected")
            
            logger.info("Updating porcion descriptions...")
            # Update porcion descriptions
            self._update_porcion_descriptions()
            logger.info("Porcion descriptions updated successfully")
            
            # Render summary page with comprehensive statistics
            return render(request, 'monitor/import_summary.html', {
                'stats': stats,
                'success': True
            })
            
        except ValueError as e:
            logger.error(f"ValueError during import: {str(e)}")
            logger.error(traceback.format_exc())
            messages.error(request, f'Error de validación: {str(e)}')
            return redirect('import_medidores')
        except MemoryError as e:
            logger.error(f"MemoryError during import: {str(e)}")
            logger.error(traceback.format_exc())
            messages.error(request, 'El archivo es demasiado grande y excede la memoria disponible. Intente con un archivo más pequeño.')
            return redirect('import_medidores')
        except Exception as e:
            logger.error(f"Unexpected error during import: {str(e)}")
            logger.error(traceback.format_exc())
            messages.error(request, f'Error al procesar el archivo: {str(e)}. Revise los logs del servidor para más detalles.')
            return redirect('import_medidores')
    
    def _process_xlsx_data(self, xlsx_file):
        """Process XLSX file with all transformations."""
        # Read Excel file
        df = pd.read_excel(xlsx_file, header=None)
        
        # Extract only columns B (index 1), D (index 3), U (index 20)
        # Note: pandas uses 0-based indexing
        if df.shape[1] < 21:  # Need at least 21 columns (0-20)
            raise ValueError('El archivo no tiene las columnas esperadas (B, D, U)')
        
        # Extract columns
        data = df[[1, 3, 20]].copy()  # B=1, D=3, U=20
        data.columns = ['numero', 'marca_original', 'porcion_original']
        
        # Remove header row if present (skip first row if it looks like a header)
        if len(data) > 0 and data.iloc[0]['numero'] and isinstance(data.iloc[0]['numero'], str):
            if not str(data.iloc[0]['numero']).replace('.', '').replace('-', '').isdigit():
                data = data.iloc[1:]
        
        # Remove empty rows
        data = data.dropna(subset=['numero', 'marca_original', 'porcion_original'])
        
        # Convert to string and clean
        data['numero'] = data['numero'].astype(str).str.strip()
        data['marca_original'] = data['marca_original'].astype(str).str.strip().str.upper()
        data['porcion_original'] = data['porcion_original'].astype(str).str.strip()
        
        # Remove empty rows again after conversion
        data = data[data['numero'] != '']
        data = data[data['marca_original'] != '']
        data = data[data['porcion_original'] != '']
        
        # Transform marcas according to rules
        marca_map = {
            'ELSTER': 'HONEYWELL',
            'HONEYWELL': 'HONEYWELL',
            'GENERAL ELECTRIC': 'TRILLIANT',
            'ITRON': 'ITRON',
            'HEXING': 'HEXING',
        }
        
        data['marca'] = data['marca_original'].map(marca_map)
        
        # Filter out unwanted marcas (ACLARA, SMART, and any others not in our map)
        data = data[data['marca'].notna()]
        
        # Normalize porciones
        data['porcion'] = data['porcion_original'].apply(self._normalize_porcion)
        
        # Filter out porciones ending in M
        data = data[~data['porcion'].str.upper().str.endswith('M')]
        
        # Filter only porciones ending in I or E
        data = data[data['porcion'].str.upper().str.endswith(('I', 'E'))]
        
        # Remove duplicates based on numero
        data = data.drop_duplicates(subset=['numero'], keep='first')
        
        # Convert to list of dicts
        processed_records = data[['numero', 'marca', 'porcion']].to_dict('records')
        
        return processed_records
    
    def _normalize_porcion(self, porcion_str):
        """
        Normalize porcion format:
        - Remove leading zeros
        - Keep last letter capitalized
        - Examples: 0401I -> 401I, 0402E -> 402E
        """
        porcion_str = str(porcion_str).strip()
        
        # Extract numbers and letter
        match = re.match(r'^0*(\d+)([IiEe])$', porcion_str)
        if not match:
            # If doesn't match expected pattern, return as is
            return porcion_str
        
        number = match.group(1)
        letter = match.group(2).upper()
        
        return f"{number}{letter}"
    
    @transaction.atomic
    def _import_data(self, processed_data):
        """
        Import medidor data and return comprehensive statistics.
        
        Returns dict with:
            - total_before: Count before import
            - total_after: Count after import
            - imported: Successfully imported
            - new_medidores: List of new medidor numbers
            - deleted_medidores: List of deleted medidor numbers
            - changed_cycle: List of dicts with numero, old_porcion, new_porcion
            - rejected_by_marca: Dict with total and breakdown by marca
        """
        # 1. Capture snapshot of existing medidores BEFORE deletion
        old_medidores = {}
        for m in Medidor.objects.select_related('porcion').all():
            old_medidores[m.numero] = {
                'marca': m.marca,
                'porcion_nombre': m.porcion.nombre if m.porcion else None
            }
        
        total_before = len(old_medidores)
        
        # 2. Create lookup of new data for comparison
        new_data_lookup = {rec['numero']: rec for rec in processed_data}
        
        # 3. Identify changes before deletion
        new_medidores_list = []
        deleted_medidores_list = []
        changed_cycle_list = []
        
        # New medidores: in new data but not in old
        for numero in new_data_lookup.keys():
            if numero not in old_medidores:
                new_medidores_list.append(numero)
        
        # Deleted medidores: in old but not in new data
        for numero in old_medidores.keys():
            if numero not in new_data_lookup:
                deleted_medidores_list.append(numero)
        
        # Changed cycle: in both but different porcion
        for numero in old_medidores.keys():
            if numero in new_data_lookup:
                old_porcion = old_medidores[numero]['porcion_nombre']
                new_porcion = new_data_lookup[numero]['porcion']
                if old_porcion != new_porcion:
                    changed_cycle_list.append({
                        'numero': numero,
                        'old_porcion': old_porcion,
                        'new_porcion': new_porcion
                    })
        
        # 4. Delete all existing medidores
        Medidor.objects.all().delete()
        
        # 5. Import new medidores with validation
        imported_count = 0
        rejected_by_marca = {}
        rejected_total = 0
        
        for record in processed_data:
            try:
                # Validate that marca exists in database (case-insensitive) - DO NOT CREATE
                marca_nombre = record['marca']
                marca = Marca.objects.filter(nombre__iexact=marca_nombre).first()
                
                if not marca:
                    # Track rejection by marca
                    rejected_by_marca[marca_nombre] = rejected_by_marca.get(marca_nombre, 0) + 1
                    rejected_total += 1
                    continue
                
                # Find or create porcion (case-insensitive)
                porcion_nombre = record['porcion']
                porcion = Porcion.objects.filter(nombre__iexact=porcion_nombre).first()
                if not porcion:
                    porcion = Porcion.objects.create(
                        nombre=porcion_nombre,
                        tipo='MASIVO'  # Default type
                    )
                
                # Create medidor with validated marca
                Medidor.objects.create(
                    numero=record['numero'],
                    marca=marca_nombre,
                    porcion=porcion
                )
                imported_count += 1
                
            except Exception as e:
                logger = logging.getLogger(__name__)
                logger.error(f"Error creating medidor {record.get('numero', 'N/A')}: {str(e)}")
                continue
        
        total_after = Medidor.objects.count()
        
        # 6. Return comprehensive statistics
        return {
            'total_before': total_before,
            'total_after': total_after,
            'imported': imported_count,
            'new_medidores': new_medidores_list,
            'deleted_medidores': deleted_medidores_list,
            'changed_cycle': changed_cycle_list,
            'rejected_by_marca': {
                'total': rejected_total,
                'by_marca': rejected_by_marca
            }
        }
    
    def _update_porcion_descriptions(self):
        """Update all porcion descriptions with meter counts by brand."""
        porciones = Porcion.objects.all()
        
        for porcion in porciones:
            # Count medidores by marca for this porcion
            counts = {
                'honeywell': porcion.medidores.filter(marca='HONEYWELL').count(),
                'trilliant': porcion.medidores.filter(marca='TRILLIANT').count(),
                'itron': porcion.medidores.filter(marca='ITRON').count(),
                'hexing': porcion.medidores.filter(marca='HEXING').count(),
            }
            
            total = sum(counts.values())
            
            if total == 0:
                porcion.descripcion = "No existen medidores AMI en esta porción"
            else:
                # Build description string with formatted numbers
                parts = []
                if counts['honeywell'] > 0:
                    parts.append(f"{counts['honeywell']:,} Honeywell".replace(',', '.'))
                if counts['itron'] > 0:
                    parts.append(f"{counts['itron']:,} Itron".replace(',', '.'))
                if counts['trilliant'] > 0:
                    parts.append(f"{counts['trilliant']:,} Trilliant".replace(',', '.'))
                if counts['hexing'] > 0:
                    parts.append(f"{counts['hexing']:,} Hexing".replace(',', '.'))
                
                # Format with proper grammar
                if len(parts) == 1:
                    marca_text = parts[0]
                elif len(parts) == 2:
                    marca_text = f"{parts[0]} y {parts[1]}"
                else:
                    marca_text = f"{', '.join(parts[:-1])} y {parts[-1]}"
                
                porcion.descripcion = f"{total:,} medidores AMI en total: {marca_text}".replace(',', '.')
            
            porcion.save()


@admin_required_method
class ImportColectoresView(View):
    """View to import medidor-collector associations from XLSX files."""
    
    def get(self, request):
        return render(request, 'monitor/import_colectores.html')
    
    def post(self, request):
        logger = logging.getLogger(__name__)
        
        if 'file' not in request.FILES:
            messages.error(request, 'No se ha seleccionado ningún archivo.')
            return redirect('import_colectores')
        
        xlsx_file = request.FILES['file']
        
        # Validate file extension
        if not xlsx_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'El archivo debe ser formato XLSX o XLS.')
            return redirect('import_colectores')
        
        try:
            # Process XLSX file
            df = pd.read_excel(xlsx_file, header=None)
            
            # Extract first two columns (Colector, Medidor)
            if df.shape[1] < 2:
                messages.error(request, 'El archivo debe tener al menos 2 columnas.')
                return redirect('import_colectores')
            
            # Extract columns
            data = df[[0, 1]].copy()  # Column 0 = Colector ID, Column 1 = Medidor número
            data.columns = ['colector_id', 'medidor_numero']
            
            # Remove header row if present
            if len(data) > 0 and data.iloc[0]['colector_id'] and isinstance(data.iloc[0]['colector_id'], str):
                if not str(data.iloc[0]['colector_id']).replace('.', '').replace('-', '').replace('_', '').isalnum():
                    data = data.iloc[1:]
            
            # Remove empty rows
            data = data.dropna(subset=['colector_id', 'medidor_numero'])
            
            # Convert to string and clean
            data['colector_id'] = data['colector_id'].astype(str).str.strip()
            data['medidor_numero'] = data['medidor_numero'].astype(str).str.strip()
            
            # Remove empty rows again after conversion
            data = data[data['colector_id'] != '']
            data = data[data['medidor_numero'] != '']
            
            # Remove duplicates (keep last occurrence for same medidor)
            data = data.drop_duplicates(subset=['medidor_numero'], keep='last')
            
            # Import associations
            stats = self._import_associations(data)
            
            return render(request, 'monitor/import_colectores_summary.html', {
                'stats': stats
            })
            
        except Exception as e:
            logger.error(f"Error processing file: {str(e)}")
            logger.error(traceback.format_exc())
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('import_colectores')
    
    @transaction.atomic
    def _import_associations(self, data):
        """
        Import medidor-collector associations and return statistics.
        
        Returns dict with:
            - created: New associations
            - updated: Changed associations
            - rejected_no_medidor: Medidor doesn't exist
            - rejected_no_colector: Colector doesn't exist
            - rejected_total: Total rejected
            - sin_colector: Medidores without association
            - total_processed: Total rows processed
        """
        created = 0
        updated = 0
        rejected_no_medidor = 0
        rejected_no_colector = 0
        
        for _, row in data.iterrows():
            colector_id = row['colector_id']
            medidor_numero = row['medidor_numero']
            
            # Check if medidor exists
            try:
                medidor = Medidor.objects.get(numero=medidor_numero)
            except Medidor.DoesNotExist:
                rejected_no_medidor += 1
                continue
            
            # Check if colector exists
            try:
                colector = Equipo.objects.get(id_equipo=colector_id)
            except Equipo.DoesNotExist:
                rejected_no_colector += 1
                continue
            
            # Check if association needs to be created or updated
            if medidor.colector is None:
                # Create new association
                medidor.colector = colector
                medidor.save()
                created += 1
            elif medidor.colector.id != colector.id:
                # Update existing association
                medidor.colector = colector
                medidor.save()
                updated += 1
            # If already associated to the same colector, do nothing
        
        # Count medidores without colector
        sin_colector = Medidor.objects.filter(colector__isnull=True).count()
        
        rejected_total = rejected_no_medidor + rejected_no_colector
        total_processed = created + updated + rejected_total
        
        return {
            'created': created,
            'updated': updated,
            'rejected_no_medidor': rejected_no_medidor,
            'rejected_no_colector': rejected_no_colector,
            'rejected_total': rejected_total,
            'sin_colector': sin_colector,
            'total_processed': total_processed,
        }
