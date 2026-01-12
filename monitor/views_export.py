import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.http import HttpResponse, HttpResponseBadRequest
from django.views import View
from django.utils import timezone
from django.db.models import Count, Q
from django.template.loader import get_template

from xhtml2pdf import pisa
from io import BytesIO
import datetime

from .models import Equipo, HistorialDisponibilidad, Marca

class ExportReportView(View):
    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('format', 'xlsx')
        
        # 1. Fetch Data (Reuse logic from ReporteView roughly)
        now = timezone.now()
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if start_date_str and end_date_str:
            try:
                current_tz = timezone.get_current_timezone()
                naive_start = datetime.datetime.strptime(start_date_str, '%Y-%m-%d')
                start_date = timezone.make_aware(naive_start, current_tz)
                naive_end = datetime.datetime.strptime(end_date_str, '%Y-%m-%d') + datetime.timedelta(days=1) - datetime.timedelta(microseconds=1)
                end_date = timezone.make_aware(naive_end, current_tz)
            except ValueError:
                start_date = now - datetime.timedelta(days=30)
                end_date = now
        else:
            start_date = now - datetime.timedelta(days=30)
            end_date = now

        # Queryset
        qs = Equipo.objects.filter(estado='ACTIVO').annotate(
            total_checks=Count('historial', filter=Q(historial__timestamp__gte=start_date, historial__timestamp__lte=end_date)),
            online_checks=Count('historial', filter=Q(historial__timestamp__gte=start_date, historial__timestamp__lte=end_date, historial__estado='ONLINE'))
        )
        
        # Filters
        query = request.GET.get('q')
        if query:
            qs = qs.filter(Q(ip__icontains=query) | Q(id_equipo__icontains=query))
            
        marca_id = request.GET.get('marca')
        if marca_id:
            qs = qs.filter(marca_id=marca_id)

        estado = request.GET.get('estado')
        if estado:
            qs = qs.filter(estado=estado)
            
        qs = qs.select_related('marca', 'tipo').order_by('id_equipo')

        # Calculate Availability
        equipos_data = []
        for equipo in qs:
            av = 0
            if equipo.total_checks > 0:
                av = round((equipo.online_checks / equipo.total_checks) * 100, 1)
            
            downtime_count = equipo.total_checks - equipo.online_checks
            
            equipos_data.append({
                'id_equipo': equipo.id_equipo,
                'ip': equipo.ip,
                'marca': equipo.marca.nombre if equipo.marca else '-',
                'tipo': equipo.tipo.nombre if equipo.tipo else '-',
                'availability': av,
                'downtime_count': downtime_count,
                'lat_lon': f"{equipo.latitud}, {equipo.longitud}"
            })

        if fmt == 'xlsx':
            return self.export_xlsx(equipos_data, start_date, end_date)
        elif fmt == 'pdf':
            return self.export_pdf(equipos_data, start_date, end_date)
        else:
            return HttpResponseBadRequest("Formato no soportado")

    def export_xlsx(self, data, start, end):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Disponibilidad"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        center_align = Alignment(horizontal="center")
        
        # Header
        headers = ["ID Equipo", "IP", "Marca", "Tipo", "Disponibilidad (%)", "Coordenadas"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Data
        for row_idx, item in enumerate(data, start=2):
            ws.cell(row=row_idx, column=1, value=item['id_equipo'])
            ws.cell(row=row_idx, column=2, value=item['ip'])
            ws.cell(row=row_idx, column=3, value=item['marca'])
            ws.cell(row=row_idx, column=4, value=item['tipo'])
            
            av_cell = ws.cell(row=row_idx, column=5, value=item['availability'])
            if item['availability'] < 75:
                av_cell.font = Font(color="DC2626", bold=True) # Red
            elif item['availability'] < 95:
                av_cell.font = Font(color="D97706", bold=True) # Orange
            else:
                av_cell.font = Font(color="059669", bold=True) # Green
            
            ws.cell(row=row_idx, column=6, value=item['lat_lon'])

        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_QAWAQ_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response

    def export_pdf(self, data, start, end):
        template_path = 'monitor/export_pdf.html'
        context = {
            'equipos': data,
            'start_date': start,
            'end_date': end,
            'generated_at': timezone.now()
        }
        
        template = get_template(template_path)
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Reporte_QAWAQ_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'

        pisa_status = pisa.CreatePDF(html, dest=response)
        
        if pisa_status.err:
            return HttpResponse('Error generating PDF', status=500)
        return response

class ExportIndividualReportView(View):
    def get(self, request, *args, **kwargs):
        from django.shortcuts import get_object_or_404
        
        # 1. Fetch Device
        equipo_code = request.GET.get('equipo_code')
        equipo = None
        
        if equipo_code:
            try:
                equipo = Equipo.objects.get(id_equipo=equipo_code)
            except Equipo.DoesNotExist:
                if str(equipo_code).isdigit():
                    try:
                        equipo = Equipo.objects.get(pk=equipo_code)
                    except Equipo.DoesNotExist:
                        pass
        
        if not equipo:
            return HttpResponseBadRequest("Equipo no encontrado o no especificado")

        # 2. Date Range
        now = timezone.now()
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if start_date_str and end_date_str:
            try:
                current_tz = timezone.get_current_timezone()
                naive_start = datetime.datetime.strptime(start_date_str, '%Y-%m-%d')
                start_date = timezone.make_aware(naive_start, current_tz)
                naive_end = datetime.datetime.strptime(end_date_str, '%Y-%m-%d') + datetime.timedelta(days=1) - datetime.timedelta(microseconds=1)
                end_date = timezone.make_aware(naive_end, current_tz)
            except ValueError:
                start_date = now - datetime.timedelta(days=30)
                end_date = now
        else:
            start_date = now - datetime.timedelta(days=30)
            end_date = now

        # 3. Calculate Stats
        total = HistorialDisponibilidad.objects.filter(
            equipo=equipo,
            timestamp__range=(start_date, end_date)
        ).count()
        
        online = HistorialDisponibilidad.objects.filter(
            equipo=equipo,
            timestamp__range=(start_date, end_date),
            estado='ONLINE'
        ).count()
        
        availability = round((online / total * 100), 2) if total > 0 else 0
        downtime_count = total - online

        # 4. Logs
        logs = HistorialDisponibilidad.objects.filter(
            equipo=equipo,
            timestamp__range=(start_date, end_date),
            estado='OFFLINE'
        ).order_by('-timestamp')

        # 5. Generate PDF
        template_path = 'monitor/export_individual_pdf.html'
        context = {
            'equipo': equipo,
            'start_date': start_date,
            'end_date': end_date,
            'generated_at': now,
            'availability': availability,
            'downtime_count': downtime_count,
            'total_checks': total,
            'logs': logs
        }
        
        template = get_template(template_path)
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Reporte_{equipo.id_equipo}_{now.strftime("%Y%m%d_%H%M")}.pdf"'

        pisa_status = pisa.CreatePDF(html, dest=response)
        
        if pisa_status.err:
            return HttpResponse('Error generating PDF', status=500)
        return response
